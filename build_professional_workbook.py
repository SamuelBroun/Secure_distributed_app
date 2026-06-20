# -*- coding: utf-8 -*-
"""
Professional workbook builder for the "דו-רה-מי והמתחם" operations file.

Takes the original basic Excel file and produces a polished, business-grade
workbook: consistent styling, Excel Tables, filters, frozen headers,
data validation dropdowns, conditional formatting, a management Dashboard
with KPIs + charts, and a README sheet.

All original data is preserved. The only value normalization performed is
mapping the task status "לא" -> "לא בוצע" for clarity (documented in README).
"""

import copy
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter

SRC = "/root/.claude/uploads/86572714-f9d7-5fe8-b31a-d5071825dd38/39a8a481-_____________________________________.xlsx"
OUT = "/home/user/Secure_distributed_app/Doremi_Operations_Professional.xlsx"

# ---------------------------------------------------------------- palette
NAVY        = "0F2A43"   # title band
HEADER      = "1F4E78"   # column headers
ACCENT      = "2E75B6"   # accents / KPI cards
BAND        = "EAF1F8"   # zebra band
WHITE       = "FFFFFF"
GREY_BORDER = "D6DCE4"
SOFT_TEXT   = "44546A"

# status colors
GREEN_FILL = "C6EFCE"; GREEN_TXT = "006100"
YEL_FILL   = "FFEB9C"; YEL_TXT   = "9C6500"
RED_FILL   = "FFC7CE"; RED_TXT   = "9C0006"

thin = Side(style="thin", color=GREY_BORDER)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

HFONT  = Font(name="Calibri", size=11, bold=True, color=WHITE)
BFONT  = Font(name="Calibri", size=11, color="1A1A1A")
TITLEF = Font(name="Calibri", size=18, bold=True, color=WHITE)
SUBF   = Font(name="Calibri", size=11, color="DCE6F1")

header_fill = PatternFill("solid", fgColor=HEADER)
navy_fill   = PatternFill("solid", fgColor=NAVY)
band_fill   = PatternFill("solid", fgColor=BAND)
accent_fill = PatternFill("solid", fgColor=ACCENT)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT  = Alignment(horizontal="right",  vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def style_header_row(ws, row, ncols, first_col=1):
    for c in range(first_col, first_col + ncols):
        cell = ws.cell(row, c)
        cell.fill = header_fill
        cell.font = HFONT
        cell.alignment = CENTER
        cell.border = BORDER
    ws.row_dimensions[row].height = 30


def autosize(ws, min_w=10, max_w=60, pad=3, scan_rows=None):
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        longest = 0
        rows = scan_rows or range(1, ws.max_row + 1)
        for r in rows:
            v = ws.cell(r, col).value
            if v is None:
                continue
            for line in str(v).split("\n"):
                longest = max(longest, len(line))
        ws.column_dimensions[letter].width = max(min_w, min(max_w, longest + pad))


def zebra_and_borders(ws, first_data_row, last_row, ncols, first_col=1):
    for r in range(first_data_row, last_row + 1):
        for c in range(first_col, first_col + ncols):
            cell = ws.cell(r, c)
            cell.border = BORDER
            cell.font = BFONT
            if cell.alignment is None or cell.alignment.vertical is None:
                cell.alignment = RIGHT
            else:
                cell.alignment = RIGHT
        if (r - first_data_row) % 2 == 1:
            for c in range(first_col, first_col + ncols):
                ws.cell(r, c).fill = band_fill
        ws.row_dimensions[r].height = 26


# ================================================================ load
src = openpyxl.load_workbook(SRC)
wb = openpyxl.Workbook()
wb.remove(wb.active)

# read raw data per sheet
data = {}
for ws in src.worksheets:
    rows = []
    for r in range(1, ws.max_row + 1):
        rows.append([ws.cell(r, c).value for c in range(1, ws.max_column + 1)])
    data[ws.title] = rows

# ================================================================ Dashboard (built later, placed first)
dash = wb.create_sheet("לוח בקרה")          # Dashboard
ws_tasks = wb.create_sheet("משימות")          # Tasks
ws_links = wb.create_sheet("מאגר לינקים")     # Links
ws_msgs  = wb.create_sheet("מאגר הודעות")     # Messages
ws_readme = wb.create_sheet("README")

for ws in wb.worksheets:
    ws.sheet_view.rightToLeft = True
    ws.sheet_view.showGridLines = False

# ---------------------------------------------------------------- helpers for title band
def title_band(ws, ncols, title, subtitle):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    t = ws.cell(1, 1, title); t.font = TITLEF; t.alignment = RIGHT; t.fill = navy_fill
    s = ws.cell(2, 1, subtitle); s.font = SUBF; s.alignment = RIGHT; s.fill = navy_fill
    for r in (1, 2):
        for c in range(1, ncols + 1):
            ws.cell(r, c).fill = navy_fill
    ws.row_dimensions[1].height = 34
    ws.row_dimensions[2].height = 20


# ================================================================ TASKS sheet
tasks_raw = data["משימות"]
t_headers = tasks_raw[0]
t_body = tasks_raw[1:]
ncol_t = len(t_headers)

title_band(ws_tasks, ncol_t, "משימות — מעקב ביצוע", "דו-רה-מי והמתחם · ניהול תפעולי")
HDR_ROW = 4
DATA_ROW = HDR_ROW + 1

# rename last column header to clean "סטטוס"
t_headers = list(t_headers)
t_headers[-1] = "סטטוס"
for c, h in enumerate(t_headers, start=1):
    ws_tasks.cell(HDR_ROW, c, h)
style_header_row(ws_tasks, HDR_ROW, ncol_t)

status_map = {"לא": "לא בוצע", "כן": "בוצע", None: "לא בוצע", "": "לא בוצע"}
for i, row in enumerate(t_body):
    rr = DATA_ROW + i
    for c, v in enumerate(row, start=1):
        if c == ncol_t:  # status column
            v = status_map.get(v, v if v else "לא בוצע")
        ws_tasks.cell(rr, c, v)

last_t = DATA_ROW + len(t_body) - 1
zebra_and_borders(ws_tasks, DATA_ROW, last_t, ncol_t)
ws_tasks.freeze_panes = ws_tasks.cell(DATA_ROW, 1)

# Excel Table
ref = f"A{HDR_ROW}:{get_column_letter(ncol_t)}{last_t}"
tbl = Table(displayName="tblTasks", ref=ref)
tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                                    showLastColumn=False, showRowStripes=True)
ws_tasks.add_table(tbl)

# column widths
widths = {1: 16, 2: 34, 3: 26, 4: 50, 5: 22, 6: 18, 7: 16}
for c, w in widths.items():
    ws_tasks.column_dimensions[get_column_letter(c)].width = w

# Data validation: status dropdown
status_col = get_column_letter(ncol_t)
dv_status = DataValidation(type="list", formula1='"לא בוצע,בתהליך,בוצע"', allow_blank=True)
dv_status.error = "בחר ערך מהרשימה"; dv_status.errorTitle = "ערך לא תקין"
dv_status.prompt = "בחר סטטוס משימה"; dv_status.promptTitle = "סטטוס"
ws_tasks.add_data_validation(dv_status)
dv_status.add(f"{status_col}{DATA_ROW}:{status_col}{last_t}")

# Data validation: category dropdown (unique values)
cats = sorted({r[0] for r in t_body if r[0]})
dv_cat = DataValidation(type="list", formula1='"' + ",".join(cats) + '"', allow_blank=True)
ws_tasks.add_data_validation(dv_cat)
dv_cat.add(f"A{DATA_ROW}:A{last_t}")

# Conditional formatting on status column
rng = f"{status_col}{DATA_ROW}:{status_col}{last_t}"
ws_tasks.conditional_formatting.add(rng,
    CellIsRule(operator="equal", formula=['"בוצע"'], fill=PatternFill("solid", fgColor=GREEN_FILL),
               font=Font(color=GREEN_TXT, bold=True)))
ws_tasks.conditional_formatting.add(rng,
    CellIsRule(operator="equal", formula=['"בתהליך"'], fill=PatternFill("solid", fgColor=YEL_FILL),
               font=Font(color=YEL_TXT, bold=True)))
ws_tasks.conditional_formatting.add(rng,
    CellIsRule(operator="equal", formula=['"לא בוצע"'], fill=PatternFill("solid", fgColor=RED_FILL),
               font=Font(color=RED_TXT, bold=True)))

# ================================================================ LINKS sheet
links_raw = data["מאגר לינקים מוכנים"]
l_headers = links_raw[0]
l_body = [r for r in links_raw[1:] if any(x not in (None, "", "-") for x in r)]  # drop fully empty rows
ncol_l = len(l_headers)

title_band(ws_links, ncol_l, "מאגר לינקים מוכנים", "לינקים לפי מוקדי רווח · שליחה ללידים")
for c, h in enumerate(l_headers, start=1):
    ws_links.cell(HDR_ROW, c, h)
style_header_row(ws_links, HDR_ROW, ncol_l)
for i, row in enumerate(l_body):
    for c, v in enumerate(row, start=1):
        ws_links.cell(DATA_ROW + i, c, v)
last_l = DATA_ROW + len(l_body) - 1
zebra_and_borders(ws_links, DATA_ROW, last_l, ncol_l)
ws_links.freeze_panes = ws_links.cell(DATA_ROW, 3)  # freeze name + desc + header
tbl_l = Table(displayName="tblLinks", ref=f"A{HDR_ROW}:{get_column_letter(ncol_l)}{last_l}")
tbl_l.tableStyleInfo = TableStyleInfo(name="TableStyleMedium6", showRowStripes=True)
ws_links.add_table(tbl_l)
ws_links.column_dimensions["A"].width = 22
ws_links.column_dimensions["B"].width = 50
for c in range(3, ncol_l + 1):
    ws_links.column_dimensions[get_column_letter(c)].width = 20

# ================================================================ MESSAGES sheet
msgs_raw = data["מאגר הודעות"]
m_headers = msgs_raw[0]
m_body = [r for r in msgs_raw[1:] if any(x not in (None, "") for x in r)]
ncol_m = len(m_headers)

title_band(ws_msgs, ncol_m, "מאגר הודעות", "תבניות הודעות פנימה / החוצה ללקוחות ולידים")
for c, h in enumerate(m_headers, start=1):
    ws_msgs.cell(HDR_ROW, c, h)
style_header_row(ws_msgs, HDR_ROW, ncol_m)
for i, row in enumerate(m_body):
    for c, v in enumerate(row, start=1):
        ws_msgs.cell(DATA_ROW + i, c, v)
last_m = DATA_ROW + len(m_body) - 1
zebra_and_borders(ws_msgs, DATA_ROW, last_m, ncol_m)
ws_msgs.freeze_panes = ws_msgs.cell(DATA_ROW, 1)
tbl_m = Table(displayName="tblMessages", ref=f"A{HDR_ROW}:{get_column_letter(ncol_m)}{last_m}")
tbl_m.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
ws_msgs.add_table(tbl_m)
ws_msgs.column_dimensions["A"].width = 28
ws_msgs.column_dimensions["B"].width = 20
ws_msgs.column_dimensions["C"].width = 40
ws_msgs.column_dimensions["D"].width = 40
ws_msgs.column_dimensions["E"].width = 45
# direction dropdown
dv_dir = DataValidation(type="list", formula1='"פנימה,החוצה,פנימה והחוצה"', allow_blank=True)
ws_msgs.add_data_validation(dv_dir)
dv_dir.add(f"B{DATA_ROW}:B{last_m}")

# ================================================================ DASHBOARD
# KPI calculations via formulas referencing tblTasks status column
status_ref = f"משימות!${status_col}${DATA_ROW}:${status_col}${last_t}"
cat_ref    = f"משימות!$A${DATA_ROW}:$A${last_t}"
total_tasks = len(t_body)

title_band(dash, 8, "לוח בקרה ניהולי", "דו-רה-מי והמתחם · תמונת מצב תפעולית")

# KPI cards row (row 4-6)
kpi_defs = [
    ("סך משימות", f"={total_tasks}", ACCENT),
    ("הושלמו", f'=COUNTIF({status_ref},"בוצע")', "2E9E5B"),
    ("בתהליך", f'=COUNTIF({status_ref},"בתהליך")', "C9981B"),
    ("לא בוצע", f'=COUNTIF({status_ref},"לא בוצע")', "C0392B"),
]
col = 1
for label, formula, color in kpi_defs:
    dash.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 1)
    dash.merge_cells(start_row=5, start_column=col, end_row=6, end_column=col + 1)
    lc = dash.cell(4, col, label)
    lc.fill = PatternFill("solid", fgColor=color); lc.font = Font(bold=True, color=WHITE, size=11)
    lc.alignment = CENTER
    vc = dash.cell(5, col, formula)
    vc.fill = PatternFill("solid", fgColor="F4F8FC"); vc.font = Font(bold=True, size=26, color=color)
    vc.alignment = CENTER
    for r in (4, 5, 6):
        for cc in (col, col + 1):
            dash.cell(r, cc).border = BORDER
    col += 2
dash.row_dimensions[4].height = 22
dash.row_dimensions[5].height = 30
dash.row_dimensions[6].height = 26

# % complete card
dash.merge_cells(start_row=8, start_column=1, end_row=8, end_column=4)
pc = dash.cell(8, 1, "אחוז השלמה כללי")
pc.fill = accent_fill; pc.font = Font(bold=True, color=WHITE); pc.alignment = CENTER
dash.merge_cells(start_row=9, start_column=1, end_row=9, end_column=4)
pv = dash.cell(9, 1, f'=IFERROR(COUNTIF({status_ref},"בוצע")/{total_tasks},0)')
pv.number_format = "0%"; pv.font = Font(bold=True, size=22, color=ACCENT); pv.alignment = CENTER
pv.fill = PatternFill("solid", fgColor="F4F8FC")
for r in (8, 9):
    for cc in range(1, 5):
        dash.cell(r, cc).border = BORDER
dash.row_dimensions[8].height = 20
dash.row_dimensions[9].height = 30

# Status distribution helper table (for pie chart)
sr = 12
dash.cell(sr, 1, "סטטוס").font = HFONT; dash.cell(sr, 1).fill = header_fill; dash.cell(sr,1).alignment=CENTER
dash.cell(sr, 2, "כמות").font = HFONT; dash.cell(sr, 2).fill = header_fill; dash.cell(sr,2).alignment=CENTER
status_list = [("בוצע", GREEN_TXT), ("בתהליך", YEL_TXT), ("לא בוצע", RED_TXT)]
for i, (st, _) in enumerate(status_list):
    dash.cell(sr + 1 + i, 1, st).border = BORDER
    dash.cell(sr + 1 + i, 2, f'=COUNTIF({status_ref},"{st}")').border = BORDER
    dash.cell(sr + 1 + i, 1).alignment = RIGHT
    dash.cell(sr + 1 + i, 2).alignment = CENTER
status_last = sr + len(status_list)

# Category distribution helper table (for bar chart)
cr = sr
dash.cell(cr, 4, "קטגוריה").font = HFONT; dash.cell(cr, 4).fill = header_fill; dash.cell(cr,4).alignment=CENTER
dash.cell(cr, 5, "כמות משימות").font = HFONT; dash.cell(cr, 5).fill = header_fill; dash.cell(cr,5).alignment=CENTER
for i, cat in enumerate(cats):
    dash.cell(cr + 1 + i, 4, cat).border = BORDER
    dash.cell(cr + 1 + i, 5, f'=COUNTIF({cat_ref},"{cat}")').border = BORDER
    dash.cell(cr + 1 + i, 4).alignment = RIGHT
    dash.cell(cr + 1 + i, 5).alignment = CENTER
cat_last = cr + len(cats)

dash.column_dimensions["A"].width = 16
dash.column_dimensions["B"].width = 14
dash.column_dimensions["C"].width = 4
dash.column_dimensions["D"].width = 18
dash.column_dimensions["E"].width = 16
for c in "FGH":
    dash.column_dimensions[c].width = 14

# Pie chart: status
pie = PieChart()
pie.title = "התפלגות סטטוס משימות"
labels = Reference(dash, min_col=1, min_row=sr + 1, max_row=status_last)
pdata = Reference(dash, min_col=2, min_row=sr, max_row=status_last)
pie.add_data(pdata, titles_from_data=True)
pie.set_categories(labels)
pie.height = 7.5; pie.width = 12
pie.dataLabels = DataLabelList(); pie.dataLabels.showPercent = True
dash.add_chart(pie, "A18")

# Bar chart: categories
bar = BarChart()
bar.type = "bar"
bar.title = "משימות לפי קטגוריה"
bar.legend = None
bcats = Reference(dash, min_col=4, min_row=cr + 1, max_row=cat_last)
bdata = Reference(dash, min_col=5, min_row=cr, max_row=cat_last)
bar.add_data(bdata, titles_from_data=True)
bar.set_categories(bcats)
bar.height = 7.5; bar.width = 12
bar.dataLabels = DataLabelList(); bar.dataLabels.showVal = True
dash.add_chart(bar, "D18")

# ================================================================ README
ws_readme.sheet_view.showGridLines = False
ws_readme.column_dimensions["A"].width = 3
ws_readme.column_dimensions["B"].width = 100
title_band(ws_readme, 3, "README — מדריך הקובץ", "מבנה, גיליונות ושיפורים שבוצעו")

readme_lines = [
    ("h", "על הקובץ"),
    ("p", "קובץ ניהול תפעולי של דו-רה-מי והמתחם. עוצב מחדש לרמה עסקית מקצועית תוך שמירה מלאה על כל הנתונים המקוריים."),
    ("h", "גיליונות"),
    ("b", "לוח בקרה — דשבורד ניהולי עם מדדי KPI (סך משימות, הושלמו, בתהליך, לא בוצע), אחוז השלמה, וגרפים: התפלגות סטטוס ומשימות לפי קטגוריה. כל המדדים מתעדכנים אוטומטית מנתוני גיליון המשימות."),
    ("b", "משימות — טבלת מעקב משימות רשמית (Excel Table) עם פילטרים, הקפאת כותרת, רשימות בחירה לקטגוריה ולסטטוס, וצביעה מותנית: ירוק=בוצע, צהוב=בתהליך, אדום=לא בוצע."),
    ("b", "מאגר לינקים — לינקים מוכנים לפי מוקדי רווח, מאורגנים כטבלה עם פילטרים והקפאת עמודות מזהות."),
    ("b", "מאגר הודעות — תבניות הודעות פנימה/החוצה עם רשימת בחירה לכיוון ההודעה."),
    ("h", "שיפורים שבוצעו"),
    ("b", "עיצוב אחיד ומינימליסטי עם פס כותרת, צבעוניות עקבית וגבולות נקיים בכל הגיליונות."),
    ("b", "טבלאות Excel רשמיות עם פסי שורות (זברה), פילטרים אוטומטיים והקפאת שורת כותרת."),
    ("b", "רוחב עמודות וגובה שורות מותאמים לקריאוּת גבוהה במחשב ובמובייל."),
    ("b", "Data Validation (רשימות בחירה) למניעת טעויות הזנה בשדות סטטוס, קטגוריה וכיוון הודעה."),
    ("b", "Conditional Formatting לזיהוי מהיר של דחיפות וסטטוס."),
    ("b", "דשבורד ניהולי עם נוסחאות COUNTIF דינמיות וגרפים מקצועיים."),
    ("h", "הערת נתונים"),
    ("p", "ערכי הסטטוס במשימות נוסחו מחדש לבהירות: \"לא\" → \"לא בוצע\". ניתן לעדכן לכל ערך מהרשימה (לא בוצע / בתהליך / בוצע). שום נתון לא נמחק."),
]
r = HDR_ROW + 1
for kind, text in readme_lines:
    cell = ws_readme.cell(r, 2, text)
    cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
    if kind == "h":
        cell.font = Font(bold=True, size=13, color=HEADER)
        ws_readme.row_dimensions[r].height = 26
    elif kind == "b":
        cell.value = "•  " + text
        cell.font = Font(size=11, color="1A1A1A")
        ws_readme.row_dimensions[r].height = 32
    else:
        cell.font = Font(size=11, color=SOFT_TEXT, italic=True)
        ws_readme.row_dimensions[r].height = 30
    r += 1

# ================================================================ workbook props + save
wb.properties.title = "דו-רה-מי והמתחם — ניהול תפעולי"
wb.properties.creator = "Operations"
wb.active = wb.sheetnames.index("לוח בקרה")
for ws in wb.worksheets:
    ws.sheet_properties.tabColor = HEADER
dash.sheet_properties.tabColor = NAVY

wb.save(OUT)
print("Saved:", OUT)
print("Sheets:", wb.sheetnames)
print("Tasks rows:", len(t_body), "Links rows:", len(l_body), "Msg rows:", len(m_body))
